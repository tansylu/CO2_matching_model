import pickle
import openpyxl

# Function to flatten a 2D list
def flatten_values_from_pickle(file_var):
    result = []
    for category in file_var.values():
        for element in category:
            result.append(element)
    return result

# Function to create inverse matching dictionary
def create_inverse_matching(categories):
    inverse_matching = {}
    for parent_category, subcategories in categories.items():
        for subcategory in subcategories:
            inverse_matching[subcategory] = parent_category
    return inverse_matching

# Function to output data to Excel
def output_to_excel(flattened_list, inverse_matching, name):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    sheet['A1'] = 'Subcategory'
    sheet['B1'] = 'Parent Category'

    # Write subcategories and parent categories
    row = 2
    for subcategory in flattened_list:
        sheet.cell(row=row, column=1, value=subcategory)
        sheet.cell(row=row, column=2, value=inverse_matching[subcategory])
        row += 1

    # Save the workbook
    workbook.save('output_{n}.xlsx'.format(n=name))

def init_processing():
    # Load the matched product data from pickle files
    ex_to_ei = pickle.load(open("product_EX_to_EI.pickle", "rb"))
    ei_to_ex = pickle.load(open("product_EI_to_EX.pickle", "rb"))

    # Create inverse matching dictionaries to determine parent category by looking at subcategory
    inverse_matching_ex_to_ei = create_inverse_matching(ex_to_ei)
    inverse_matching_ei_to_ex = create_inverse_matching(ei_to_ex)

    # Obtain flattened value lists for both dictionaries
    ex_to_ei_flat = flatten_values_from_pickle(ex_to_ei)
    ei_to_ex_flat = flatten_values_from_pickle(ei_to_ex)

    # Output the flattened lists to Excel with parent categories
    output_to_excel(ex_to_ei_flat, inverse_matching_ex_to_ei, "ex_to_ei")
    output_to_excel(ei_to_ex_flat, inverse_matching_ei_to_ex, "ei_to_ex")
